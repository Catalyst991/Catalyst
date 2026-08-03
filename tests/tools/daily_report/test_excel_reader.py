import datetime

import openpyxl
import pytest

from catalyst.tools.daily_report.excel_reader import MissingColumnError, MissingSheetError, read_comments

COLUMNS = ["Date", "User name", "Comment", "Link", "Country", "No. of Followers", "النبرة"]


def make_workbook(path, rows, sheet_name="Users"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(COLUMNS)
    for row in rows:
        ws.append(row)
    wb.save(path)
    return path


@pytest.fixture
def workbook_path(tmp_path):
    return tmp_path / "workbook.xlsx"


def test_reads_a_single_comment_from_the_users_sheet(workbook_path):
    make_workbook(
        workbook_path,
        rows=[
            [datetime.date(2026, 7, 27), "Design Studio", "hello world", "https://x.com/1", "KSA", 1296, "محايد"],
        ],
    )

    comments = read_comments(workbook_path)

    assert len(comments) == 1
    comment = comments[0]
    assert comment.date == datetime.date(2026, 7, 27)
    assert comment.user_name == "Design Studio"
    assert comment.comment == "hello world"
    assert comment.link == "https://x.com/1"
    assert comment.country == "KSA"
    assert comment.follower_count == 1296
    assert comment.tone == "محايد"


def test_ignores_the_official_sheet_and_reads_only_users(workbook_path):
    wb = openpyxl.Workbook()
    users_ws = wb.active
    users_ws.title = "Users"
    users_ws.append(COLUMNS)
    users_ws.append([datetime.date(2026, 7, 27), "Design Studio", "hello world", "https://x.com/1", "KSA", 1296, "محايد"])
    official_ws = wb.create_sheet("Official")
    official_ws.append(["Should", "Not", "Be", "Read"])
    official_ws.append(["a", "b", "c", "d"])
    wb.save(workbook_path)

    comments = read_comments(workbook_path)

    assert len(comments) == 1
    assert comments[0].user_name == "Design Studio"


def test_raises_missing_sheet_error_when_no_users_sheet(workbook_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Official"
    wb.save(workbook_path)

    with pytest.raises(MissingSheetError) as exc_info:
        read_comments(workbook_path)

    assert str(exc_info.value) == (
        "This file doesn't have a 'Users' sheet with the expected data — please check the file and try again."
    )


def test_raises_missing_column_error_when_a_required_column_is_absent(workbook_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    columns_without_comment = [c for c in COLUMNS if c != "Comment"]
    ws.append(columns_without_comment)
    wb.save(workbook_path)

    with pytest.raises(MissingColumnError) as exc_info:
        read_comments(workbook_path)

    assert str(exc_info.value) == (
        "This file is missing the 'Comment' column — please check the file and try again."
    )


def test_skips_trailing_blank_rows(workbook_path):
    make_workbook(
        workbook_path,
        rows=[
            [datetime.date(2026, 7, 27), "Design Studio", "hello world", "https://x.com/1", "KSA", 1296, "محايد"],
            ["", "", "", "", "", "", ""],
        ],
    )

    comments = read_comments(workbook_path)

    assert len(comments) == 1
