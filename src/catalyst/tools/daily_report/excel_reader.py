from dataclasses import dataclass
from datetime import date

import openpyxl

SHEET_NAME = "Users"
REQUIRED_COLUMNS = ["Date", "User name", "Comment", "Link", "Country", "No. of Followers", "النبرة"]


class ExcelValidationError(Exception):
    """Base for validation failures with a plain-language, user-facing message."""


class MissingSheetError(ExcelValidationError):
    def __init__(self):
        super().__init__(
            "This file doesn't have a 'Users' sheet with the expected data — please check the file and try again."
        )


class MissingColumnError(ExcelValidationError):
    def __init__(self, column):
        self.column = column
        super().__init__(f"This file is missing the '{column}' column — please check the file and try again.")


@dataclass
class Comment:
    date: date
    user_name: str
    comment: str
    link: str
    country: str
    follower_count: int
    tone: str


def read_comments(path) -> list[Comment]:
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise MissingSheetError()
    ws = wb[SHEET_NAME]

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    for column in REQUIRED_COLUMNS:
        if column not in header:
            raise MissingColumnError(column)

    comments = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in row):
            continue
        raw_date, user_name, comment, link, country, follower_count, tone = row
        comments.append(
            Comment(
                date=raw_date.date() if hasattr(raw_date, "date") else raw_date,
                user_name=user_name,
                comment=comment,
                link=link,
                country=country,
                follower_count=follower_count,
                tone=tone,
            )
        )
    return comments
