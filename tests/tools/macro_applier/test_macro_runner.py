import shutil
from pathlib import Path

import openpyxl
import pytest

from catalyst.tools.macro_applier.macro_runner import MacroSession
from catalyst.tools.macro_applier.tasks import (
    DAILY_REPORT_FORMATTING,
    MAIN_TWITTER,
    SOCIAL_FILE_ARRANGEMENT,
    TALKWALKER_TRADITIONAL,
)

RAW_EXPORT_SAMPLE = Path(__file__).parent / "fixtures" / "raw_export_sample.xlsx"


def test_applies_the_real_macro_and_saves_the_target_file_in_place(tmp_path):
    target_path = tmp_path / "target.xlsx"
    shutil.copy(RAW_EXPORT_SAMPLE, target_path)

    session = MacroSession(target_path)
    try:
        session.apply(MAIN_TWITTER.path, MAIN_TWITTER.macro_name)
        assert session.target_workbook.Saved
    finally:
        session.close()

    wb = openpyxl.load_workbook(target_path)
    assert "Users" in wb.sheetnames


def test_close_is_safe_to_call_twice(tmp_path):
    target_path = tmp_path / "target.xlsx"
    shutil.copy(RAW_EXPORT_SAMPLE, target_path)

    session = MacroSession(target_path)
    session.apply(MAIN_TWITTER.path, MAIN_TWITTER.macro_name)
    session.close()
    session.close()


def test_a_macro_with_no_dialogs_returns_its_status_and_message_instead_of_popping_a_box(
    tmp_path,
):
    target_path = tmp_path / "target.xlsx"
    shutil.copy(RAW_EXPORT_SAMPLE, target_path)

    session = MacroSession(target_path)
    try:
        result = session.apply(DAILY_REPORT_FORMATTING.path, DAILY_REPORT_FORMATTING.macro_name)
    finally:
        session.close()

    assert result == ("DONE", "Done.")


def test_a_confirmation_macro_previews_without_changing_anything_then_applies_on_confirm(
    tmp_path,
):
    target_path = tmp_path / "target.xlsx"
    shutil.copy(RAW_EXPORT_SAMPLE, target_path)

    session = MacroSession(target_path)
    try:
        status, message = session.apply(
            TALKWALKER_TRADITIONAL.path, TALKWALKER_TRADITIONAL.macro_name, False
        )
        assert status == "CONFIRM"
        assert "Continue?" in message
        session.target_workbook.Save()

        wb = openpyxl.load_workbook(target_path, read_only=True)
        assert "Budget" not in wb.sheetnames
        assert wb.active["A1"].value == "url"
        wb.close()

        status, message = session.apply(
            TALKWALKER_TRADITIONAL.path, TALKWALKER_TRADITIONAL.macro_name, True
        )
        assert status == "DONE"
        assert "converted successfully" in message
    finally:
        session.close()

    wb = openpyxl.load_workbook(target_path)
    assert "Budget" in wb.sheetnames


@pytest.mark.parametrize(
    "first,second",
    [
        (MAIN_TWITTER, SOCIAL_FILE_ARRANGEMENT),
        (SOCIAL_FILE_ARRANGEMENT, MAIN_TWITTER),
    ],
)
def test_a_multi_macro_tasks_two_buttons_work_correctly_in_either_click_order(
    tmp_path, first, second
):
    def succeeded(result):
        return result is None or (isinstance(result, tuple) and result[0] == "DONE")

    target_path = tmp_path / "target.xlsx"
    shutil.copy(RAW_EXPORT_SAMPLE, target_path)

    session = MacroSession(target_path)
    try:
        result1 = session.apply(first.path, first.macro_name)
        result2 = session.apply(second.path, second.macro_name)
    finally:
        session.close()

    assert succeeded(result1)
    assert succeeded(result2)


def test_a_graceful_macro_failure_leaves_the_target_file_on_disk_unchanged(tmp_path):
    target_path = tmp_path / "target.xlsx"
    openpyxl.Workbook().save(target_path)
    original_bytes = target_path.read_bytes()

    session = MacroSession(target_path)
    try:
        status, message = session.apply(
            TALKWALKER_TRADITIONAL.path, TALKWALKER_TRADITIONAL.macro_name, True
        )
        assert status == "ERROR"
        assert "does not contain Raw-style data" in message

        # The session must still be usable afterward - the target workbook
        # was reloaded, not the whole session torn down.
        assert session.target_workbook is not None
        assert session.excel is not None
    finally:
        session.close()

    assert target_path.read_bytes() == original_bytes


def test_a_raised_exception_tears_down_the_session_with_no_orphaned_excel(tmp_path):
    target_path = tmp_path / "target.xlsx"
    shutil.copy(RAW_EXPORT_SAMPLE, target_path)

    session = MacroSession(target_path)
    with pytest.raises(Exception):
        session.apply(MAIN_TWITTER.path, "NoSuchMacroSubName")

    assert session.excel is None
    assert session.target_workbook is None


def test_a_locked_target_file_is_rejected_with_a_clean_message_before_touching_excel(tmp_path):
    target_path = tmp_path / "target.xlsx"
    shutil.copy(RAW_EXPORT_SAMPLE, target_path)

    blocker = MacroSession(target_path)
    try:
        with pytest.raises(RuntimeError, match="currently open in another program"):
            MacroSession(target_path)
    finally:
        blocker.close()


def test_a_missing_macro_file_shows_excels_own_clean_message_not_a_raw_com_error(tmp_path):
    target_path = tmp_path / "target.xlsx"
    shutil.copy(RAW_EXPORT_SAMPLE, target_path)

    session = MacroSession(target_path)
    with pytest.raises(RuntimeError) as excinfo:
        session.apply(tmp_path / "does_not_exist.xlsm", "Whatever")

    message = str(excinfo.value)
    assert "moved, renamed or deleted" in message
    assert "com_error" not in message
    assert session.excel is None


def test_applying_a_macro_to_an_xlsm_target_preserves_its_own_existing_macros(tmp_path):
    # openpyxl's keep_vba archive emits a harmless PytestUnraisableExceptionWarning
    # on GC (an upstream openpyxl quirk, not caused by anything here).
    target_path = tmp_path / "target.xlsm"
    openpyxl.load_workbook(MAIN_TWITTER.path, keep_vba=True).save(target_path)

    session = MacroSession(target_path)
    try:
        result = session.apply(DAILY_REPORT_FORMATTING.path, DAILY_REPORT_FORMATTING.macro_name)
    finally:
        session.close()

    assert result == ("DONE", "Done.")
    reloaded = openpyxl.load_workbook(target_path, keep_vba=True)
    assert reloaded.vba_archive is not None
